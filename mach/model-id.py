#!/bin/python
import openpyxl

excel = openpyxl.load_workbook(filename="./MACH_emaili.xlsx")

sheet = excel.active


import json
cache = {}


with open("./parsed-config.json") as file:    
    configs = json.load(file)    
    for c in configs:
        machId = c["machId"]
        timestamp = c["timestamp"]
        model = c["config"]["system"]["vehicle"]["model"]
        # print(machId, timestamp, model)
        if machId not in cache:
            cache[machId] = [machId, timestamp, model]
        elif timestamp > cache[machId][1]:
            cache[machId] = [machId, timestamp, model]


    for c in cache:
        mach = cache[c]
        # print(mach[0], mach[2])

print("Found", len(cache), "MACHs in configs")
print("Found", (sheet.max_row), "rows in XLSX")
# ignore header
for i in range(2,sheet.max_row+1):
    machId = sheet.cell(row=i, column=4).value
    if machId in cache:
        sheet.cell(row=i, column=6).value = cache[machId][2]
    else:
        sheet.cell(row=i, column=6).value = "Unknown"


excel.save(filename="parsed.xlsx")